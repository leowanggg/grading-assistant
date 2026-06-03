"""
Scoring module — score aggregation and .xlsx export.

Scoring rules:
  Each sub-question = 5 pts  (程序=3, 结果=2)
  Each major question = 5 × 5 = 25 pts
  Total = 4 × 25 = 100 pts
"""

from __future__ import annotations

import os
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from docx_parser import StudentSubmission, StudentInfo

# ---------------------------------------------------------------------------
# Score configuration
# ---------------------------------------------------------------------------

SCORE_CONFIG = {
    "program_max": 3.0,   # 程序 score max
    "result_max": 2.0,    # 结果 score max
    "sub_max": 5.0,       # total per sub-question
    "major_count": 4,     # number of major questions
    "sub_per_major": 5,   # sub-questions per major
}

# ---------------------------------------------------------------------------
# Score computation helpers
# ---------------------------------------------------------------------------

def sub_score(scores: Dict[str, float], major: int, sub: int) -> float:
    """Total for one sub-question (程序 + 结果)."""
    prog = scores.get(f"{major}-{sub}-程序", 0.0)
    res  = scores.get(f"{major}-{sub}-结果", 0.0)
    return min(prog + res, SCORE_CONFIG["sub_max"])


def major_score(scores: Dict[str, float], major: int) -> float:
    """Total for one major question (sum of 5 sub-questions)."""
    total = sum(sub_score(scores, major, s) for s in range(1, 6))
    return min(total, SCORE_CONFIG["sub_per_major"] * SCORE_CONFIG["sub_max"])


def total_score(scores: Dict[str, float]) -> float:
    """Overall total across all 4 major questions."""
    total = sum(major_score(scores, m) for m in range(1, 5))
    return min(total, SCORE_CONFIG["major_count"] * SCORE_CONFIG["sub_per_major"] * SCORE_CONFIG["sub_max"])


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

# Column styling
HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, row: int, max_col: int):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data(ws, row: int, max_col: int, score_cols: set):
    """Apply data styling — score columns centered, info columns left."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        if col in score_cols:
            cell.alignment = CENTER
            cell.number_format = "0.0"
        else:
            cell.alignment = Alignment(vertical="center")


def export_to_xlsx(
    submissions: List[StudentSubmission],
    all_scores: Dict[str, Dict[str, float]],
    output_path: str,
) -> str:
    """
    Generate grading spreadsheet.

    Columns:
      A: 学号  B: 姓名  C: 专业班级
      D-H: Q1(1-5)程序  I-M: Q1(1-5)结果
      N-R: Q2(1-5)程序  S-W: Q2(1-5)结果
      X-AB: Q3(1-5)程序  AC-AG: Q3(1-5)结果
      AH-AL: Q4(1-5)程序  AM-AQ: Q4(1-5)结果
      AR: 第一题总分  AS: 第二题总分  AT: 第三题总分  AU: 第四题总分  AV: 总分
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "批改成绩"

    # --- Build headers ---
    headers = ["学号", "姓名", "专业班级"]
    for q in range(1, 5):
        for sub in range(1, 6):
            headers.append(f"Q{q}({sub})程序")
    for q in range(1, 5):
        for sub in range(1, 6):
            headers.append(f"Q{q}({sub})结果")
    for q in range(1, 5):
        headers.append(f"第{q}题总分")
    headers.append("总分")

    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "D2"

    # Pre-compute which columns are scores
    score_columns = set(range(4, len(headers) + 1))  # cols D onward

    # --- Data rows ---
    for sub_info in submissions:
        scores = all_scores.get(sub_info.filename, {})
        row = [
            sub_info.student_info.student_id or "",
            sub_info.student_info.name or "",
            sub_info.student_info.class_name or "",
        ]
        # 程序 scores
        for q in range(1, 5):
            for s in range(1, 6):
                row.append(scores.get(f"{q}-{s}-程序", 0.0))
        # 结果 scores
        for q in range(1, 5):
            for s in range(1, 6):
                row.append(scores.get(f"{q}-{s}-结果", 0.0))
        # Major question totals
        for q in range(1, 5):
            row.append(major_score(scores, q))
        # Grand total
        row.append(total_score(scores))

        ws.append(row)
        _style_data(ws, ws.max_row, len(headers), score_columns)

    # --- Column widths ---
    col_widths = {1: 14, 2: 10, 3: 18}
    for col_idx in range(4, len(headers) + 1):
        col_widths[col_idx] = 10
    # Wider for total columns
    for q in range(1, 5):
        col_widths[len(headers) - 5 + q] = 12
    col_widths[len(headers)] = 10

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Highlight total columns ---
    total_start = len(headers) - 4  # 第1题总分 column
    for row in range(1, ws.max_row + 1):
        for col in range(total_start, total_start + 5):
            cell = ws.cell(row=row, column=col)
            if row == 1:
                cell.fill = TOTAL_FILL
            else:
                cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                cell.font = Font(bold=True)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Quick test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Dummy test
    dummy_scores = {}
    for q in range(1, 5):
        for s in range(1, 6):
            dummy_scores[f"{q}-{s}-程序"] = 2.5
            dummy_scores[f"{q}-{s}-结果"] = 1.5
    print(f"Total (dummy): {total_score(dummy_scores)}")  # expect 80.0
    print(f"Major 1: {major_score(dummy_scores, 1)}")     # expect 20.0
    print(f"Sub 1-1: {sub_score(dummy_scores, 1, 1)}")   # expect 4.0
